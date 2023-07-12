from tkinter import *
import openpyxl
from datetime import datetime

#GET ALL INPUTS
def GetAll():
    global Collect,Company,Floor,ReplaceLost,FRDoor,RearDoor,BShed,CarparkB,ExitBarrier,Fst,Trd,Fth,Eth,Collect,Lost,wb,window
    FirstName = FNameInput()
    LastName = SNameInput()
    FobID = IDInput()
    CarReg = RegInput()
    ReturnDate = ReturnInput()
    Fire = FireInput()
    Date = getDate()
    CarSpace = SpaceInput()
    FloorCompany = str(Floor)+" "+str(Company)
    ws = wb["FobAccess"]
    FobNum = GetFobNum(ws)
    new_row = (Date,FobNum,FirstName,LastName,Fire,FloorCompany,ReplaceLost,FRDoor,RearDoor,CarparkB,ExitBarrier,BShed,Fst,Trd,Fth,Eth,FobID,Collect,CarReg,CarSpace,Lost,ReturnDate)
    ws.append(new_row)
    wb.save("FobAccess.xlsx")
    window.destroy()

#GET FOB NUMBER
def GetFobNum(ws):
    cell = "B"+str(ws.max_row)
    FobNum = ws[cell].value
    FobNum = FobNum + 1
    return FobNum

#GET DATE
def getDate():
    Date = datetime.today().strftime('%d/%m/%Y')
    return Date
    
#CHECKBOX ACCESS
def FRDoors():
    global FRDoor
    if var1.get() == 1:
        FRDoor = 1
    else:
        FRDoor = 0
def RearDoors():
    global RearDoor
    if var5.get() == 1:
        RearDoor = 1
    else:
        RearDoor = 0
def BikeShed():
    global BShed
    if var2.get() == 1:
        BShed = 1
    else:
        BShed = 0
def Carpark():
    global CarparkB
    if var3.get() == 1:
        CarparkB = 1
        textboxReg.pack()
        textboxSpace.pack()
    else:
        CarparkB = 0
        textboxReg.pack_forget()
        textboxSpace.pack_forget()
def ExitBarriers():
    global ExitBarrier
    if var6.get() == 1:
        ExitBarrier = 1
    else:
        ExitBarrier = 0
def FstF():
    global Fst
    if var7.get() == 1:
        Fst = 1
    else:
        Fst = 0
def TrdF():
    global Trd
    if var8.get() == 1:
        Trd = 1
    else:
        Trd = 0
def FthF():
    global Fth
    if var9.get() == 1:
        Fth = 1
    else:
        Fth = 0
def EthF():
    global Eth
    if var4.get() == 1:
        Eth = 1
    else:
        Eth = 0

#GET TEXT INPUTS
def FNameInput():
    inp = textboxFName.get()
    FirstName = inp.upper()
    return FirstName
def temp_textFName(e):
   textboxFName.delete(0,"end")

def SNameInput():
    inp = textboxSName.get()
    LastName = inp.upper()
    return LastName
def temp_textSName(e):
   textboxSName.delete(0,"end")

def FireInput():
    inp = textboxFire.get()
    Fire = inp.upper()
    if Fire == "FIRST AIDERS & FIRE MARSHALS":
        Fire = ""
    return Fire
def temp_textFire(e):
   textboxFire.delete(0,"end")

def IDInput():
    inp = textboxID.get()
    FobID = inp
    return FobID
def temp_textID(e):
   textboxID.delete(0,"end")

def RegInput():
    inp = textboxReg.get()
    CarReg = inp
    if CarReg == "Car Registration":
        CarReg = ""
    return CarReg
def temp_textReg(e):
   textboxReg.delete(0,"end")

def SpaceInput():
    inp = textboxSpace.get()
    CarSpace = inp
    if CarSpace == "Carpark space number":
        CarSpace = ""
    return CarSpace
def temp_textSpace(e):
   textboxSpace.delete(0,"end")

def ReturnInput():
    inp = textboxReturn.get()
    ReturnDate = inp
    if ReturnDate == "Return Date":
        ReturnDate = ""
    return ReturnDate
def temp_textReturn(e):
   textboxReturn.delete(0,"end")

#GET COMPANY FROM DROPDOWN
def GetCompany(selection):
    global Company
    Company = selection

#GET FLOOR FROM DROPDOWN
def GetFloor(selection):
    global Floor
    Floor = selection

#GET LOST FROM DROPDOWN
def GetLost(selection):
    global ReplaceLost
    if selection != "None":
        ReplaceLost = selection

#GET COLLECT FROM DROPDOWN
def CollectReturn(selection):
    global Collect, Lost
    if selection == "Collected":
        Collect = 1
        Lost = 0
        textboxReturn.delete(0,"end")
        textboxReturn.pack_forget()
    elif selection == "Returned":
        Lost = 0
        Collect = 0
        if str(textboxReturn.get()) != "Return Date":
            textboxReturn.insert(0,"Return Date")
        textboxReturn.pack()
    elif selection == "Lost":
        Lost = 1
        Collect = 0
        textboxReturn.delete(0,"end")
        textboxReturn.pack_forget()

#OPEN EXCEL
wb = openpyxl.load_workbook("FobAccess.xlsx")
#CREATE GLOBAL VARIABLES
Collect,Company,Floor,ReplaceLost,FRDoor,RearDoor,BShed,CarparkB,ExitBarrier,Fst,Trd,Fth,Eth,Lost = 0,"","","",0,0,0,0,0,0,0,0,0,0
#FORMAT WINDOW
window = Tk()
window.title('Fob Stock')
window.geometry('400x600')

#CREATE CHECKLIST VARIABLES
var1 = IntVar()
var2 = IntVar()
var3 = IntVar()
var4 = IntVar()
var5 = IntVar()
var6 = IntVar()
var7 = IntVar()
var8 = IntVar()
var9 = IntVar()

#CREATE TEXT BOXES
textboxFName = Entry(window, bg="white", width=50, borderwidth=2)
textboxFName.insert(0, "First Name")
textboxFName.pack()
textboxFName.bind("<FocusIn>", temp_textFName)

textboxSName = Entry(window, bg="white", width=50, borderwidth=2)
textboxSName.insert(0, "Last Name")
textboxSName.pack()
textboxSName.bind("<FocusIn>", temp_textSName)

textboxFire = Entry(window, bg="white", width=50, borderwidth=2)
textboxFire.insert(0, "First Aiders & Fire Marshals")
textboxFire.pack()
textboxFire.bind("<FocusIn>", temp_textFire)

textboxID = Entry(window, bg="white", width=50, borderwidth=2)
textboxID.insert(0, "Fob ID")
textboxID.pack()
textboxID.bind("<FocusIn>", temp_textID)

textboxReg = Entry(window, bg="white", width=50, borderwidth=2)
textboxReg.insert(0, "Car Registration")
textboxReg.bind("<FocusIn>", temp_textReg)

textboxSpace = Entry(window, bg="white", width=50, borderwidth=2)
textboxSpace.insert(0, "Carpark space number")
textboxSpace.bind("<FocusIn>", temp_textSpace)

textboxReturn = Entry(window, bg="white", width=50, borderwidth=2)
textboxReturn.insert(0, "Return Date")
textboxReturn.bind("<FocusIn>", temp_textReturn)

#CREATE CHECKLIST ACCESS
c1 = Checkbutton(window, text="Front door", variable=var1, onvalue=1, offvalue=0, command=FRDoors)
c1.pack()
c5 = Checkbutton(window, text="Rear door", variable=var5, onvalue=1, offvalue=0, command=RearDoors)
c5.pack()
c3 = Checkbutton(window, text="Entry barrier", variable=var3, onvalue=1, offvalue=0, command=Carpark)
c3.pack()
c6 = Checkbutton(window, text="Exit barrier", variable=var6, onvalue=1, offvalue=0, command=ExitBarriers)
c6.pack()
c2 = Checkbutton(window, text="Bike shed", variable=var2, onvalue=1, offvalue=0, command=BikeShed)
c2.pack()
c7 = Checkbutton(window, text="1st", variable=var7, onvalue=1, offvalue=0, command=FstF)
c7.pack()
c8 = Checkbutton(window, text="3rd", variable=var8, onvalue=1, offvalue=0, command=TrdF)
c8.pack()
c9 = Checkbutton(window, text="4th", variable=var9, onvalue=1, offvalue=0, command=FthF)
c9.pack()
c4 = Checkbutton(window, text="8th", variable=var4, onvalue=1, offvalue=0, command=EthF)
c4.pack()

#CREATE DROPDOWN
menu= StringVar()
menu.set("Select Company")
company_file = open("company.txt", "r")
data = company_file.read()
companylist = data.split("\n")
company_file.close()
drop = OptionMenu(window, menu, *companylist, command = GetCompany)
drop.pack()

#CREATE DROPDOWN
menu= StringVar()
menu.set("Select Floor")
floorlist = ["G","1","2","3","4","5","6","7","8","9","10"]
drop = OptionMenu(window, menu, *floorlist, command = GetFloor)
drop.pack()

#CREATE DROPDOWN
menu= StringVar()
menu.set("New/Replacement/Faulty")
choicelist = ["New","Replacement for lost","Faulty","None"]
drop = OptionMenu(window, menu, *choicelist, command = GetLost)
drop.pack()

#CREATE DROPDOWN
menu= StringVar()
menu.set("Collected/Returned/Lost")
choiceslist = ["Collected","Returned","Lost","None"]
drop = OptionMenu(window, menu, *choiceslist, command = CollectReturn)
drop.pack()

#CREATE END BUTTON
ButtonEnd = Button(window, text="Enter", command=GetAll)
ButtonEnd.pack()  


window.mainloop()
