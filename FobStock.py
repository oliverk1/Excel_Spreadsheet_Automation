from tkinter import *
import openpyxl
from datetime import datetime

#GET ALL INPUTS
def GetAll():
    global Company, Floor, Issued, Returned, wb, window
    FirstName = FNameInput()
    LastName = SNameInput()
    FobID = IDInput()
    Date = getDate()
    ws = wb["FobStock"]
    fobnumafter, fobnumbefore = getNumBeforeAfter(ws)
    UserNum = getUserNum(ws)
    ReturnedFobNum = ""
    if Returned == 1:
        ReturnedFobNum = GetReturnedFobNum(ws,FobID)
    new_row = (Date,fobnumbefore,Issued,Returned,fobnumafter,UserNum,ReturnedFobNum,LastName,FirstName,Floor,Company,FobID)
    ws.append(new_row)
    wb.save("FobStock.xlsx")
    window.destroy()

#GET FOB NUMBERS
def getNumBeforeAfter(ws):
    global Issued, Returned
    cell = "E"+str(ws.max_row)
    fobnumbefore = ws[cell].value
    fobnumafter = fobnumbefore
    if Issued == 1 and Returned == 0:
        fobnumafter = int(fobnumbefore) - 1
    elif Issued == 0 and Returned == 1:
        fobnumafter = int(fobnumbefore) + 1
    return fobnumafter, fobnumbefore

#FIND PREVIOUS FOB NUM
def GetReturnedFobNum(ws, FobID):
    CellLocation = "None"
    for i in range(ws.max_row):
        num = i+1
        cell = "L"+str(num)
        currentID = ws[cell].value
        if str(currentID) == str(FobID):
            CellLocation = num
    if CellLocation != "None":
        cell = "F"+str(CellLocation)
        ReturnedFobNum = ws[cell].value
    else:
        ReturnedFobNum = "N/A"
    return ReturnedFobNum

#NEXT USER NUM
def getUserNum(ws):
    cell = "F"+str(ws.max_row)
    UserNum = ws[cell].value
    UserNum = int(UserNum)+1
    return UserNum

#GET DATE
def getDate():
    Date = datetime.today().strftime('%d/%m/%Y')
    return Date
    
#CHECKBOX ACCESS
def IsIssued():
    global Issued
    if var1.get() == 1:
        Issued = 1
    else:
        Issued = 0
def IsReturned():
    global Returned
    if var2.get() == 1:
        Returned = 1
    else:
        Returned = 0

#GET TEXT INPUTS
def FNameInput():
    inp = textboxFName.get()
    FirstName = inp.upper()
    return FirstName
def temp_textFName(e):
   textboxFName.delete(0,"end")

def SNameInput():
    inp = textboxSName.get()
    LastName = inp.upper()
    return LastName
def temp_textSName(e):
   textboxSName.delete(0,"end")

def IDInput():
    inp = textboxID.get()
    FobID = inp
    return FobID
def temp_textID(e):
   textboxID.delete(0,"end")

#GET COMPANY FROM DROPDOWN
def GetCompany(selection):
    global Company
    Company = selection

#GET FLOOR FROM DROPDOWN
def GetFloor(selection):
    global Floor
    Floor = selection

#OPEN EXCEL
wb = openpyxl.load_workbook("FobStock.xlsx")
#CREATE GLOBAL VARIABLES
Company, Floor, Issued, Returned = "","",0,0
#FORMAT WINDOW
window = Tk()
window.title('Fob Stock')
window.geometry('200x200')

#CREATE CHECKLIST VARIABLES
var1 = IntVar()
var2 = IntVar()

#CREATE TEXT BOXES
textboxFName = Entry(window, bg="white", width=50, borderwidth=2)
textboxFName.insert(0, "First Name")
textboxFName.pack()
textboxFName.bind("<FocusIn>", temp_textFName)

textboxSName = Entry(window, bg="white", width=50, borderwidth=2)
textboxSName.insert(0, "Last Name")
textboxSName.pack()
textboxSName.bind("<FocusIn>", temp_textSName)

textboxID = Entry(window, bg="white", width=50, borderwidth=2)
textboxID.insert(0, "Fob ID")
textboxID.pack()
textboxID.bind("<FocusIn>", temp_textID)

#CREATE CHECKLIST ACCESS
c1 = Checkbutton(window, text="Issued", variable=var1, onvalue=1, offvalue=0, command=IsIssued)
c1.pack()
c2 = Checkbutton(window, text="Returned", variable=var2, onvalue=1, offvalue=0, command=IsReturned)
c2.pack()

#CREATE DROPDOWN
menu= StringVar()
menu.set("Select Company")
company_file = open("company.txt", "r")
data = company_file.read()
companylist = data.split("\n")
company_file.close()
drop= OptionMenu(window, menu, *companylist, command = GetCompany)
drop.pack()

#CREATE DROPDOWN
menu= StringVar()
menu.set("Select Floor")
floorlist = ["G","1","2","3","4","5","6","7","8","9","10"]
drop= OptionMenu(window, menu, *floorlist, command = GetFloor)
drop.pack()

#CREATE END BUTTON
ButtonEnd = Button(window, text="Enter", command=GetAll)
ButtonEnd.pack()  


window.mainloop()
